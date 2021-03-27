import os
import csv
from openpyxl import load_workbook
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

import support
import ofxparse



def retornarListarArqsOFX(pasta):
    caminhos = [os.path.join(pasta, nome) for nome in os.listdir(pasta)]
    arquivos = [arq for arq in caminhos if os.path.isfile(arq)]
    listaOFX = [arq for arq in arquivos if arq.lower().endswith(".ofx")]
    return listaOFX

def inserirDadosCSV(listaOFX):
    for arq in listaOFX:
        ofx = ofxparse.OfxParser.parse(support.open_file(arq))
        account = ofx.account
        statement = account.statement
        institution = account.institution

        if account.routing_number != '':
            banco = retornarNomeBanco(account.routing_number)
        elif institution.fid != '':
            banco = retornarNomeBanco(institution.fid)
        conta = retornarNomeConta(account.account_id)
        print(banco,conta)
        rows = []
        for transaction in statement.transactions:
            # print(date_time)
            rows.append(('', '', banco + " " + conta, '',
                         transaction.memo, '', '', transaction.date,
                         transaction.amount, '', 'BRL', ''))

        if os.path.exists('dados.csv'):
            with open('dados.csv', 'a', newline='', encoding='utf-8') as arq:
                dados = csv.writer(arq, lineterminator='\n', delimiter=';')
                dados.writerows(rows)
        else:
            with open('dados.csv', 'w') as arq:
                dados = csv.writer(arq, lineterminator='\n', delimiter=';')
                dados.writerow(
                    ['Nome', 'Saldo atual', 'Conta', 'Transferências', 'Descrição', 'Beneficiário', 'Categoria',
                     'Data', 'Tempo', 'Valor', 'Câmbio', 'Número do cheque', 'Saldo'])
                dados.writerows(rows)

def retornarNomeBanco(routing_number):
    if routing_number == '1':
        return 'BB'
    elif routing_number == '0104':
        return 'CX'
    elif routing_number == '0260':
        return 'Nu'
    else:
        return ''


def retornarNomeConta(numeroConta):
    if (numeroConta == '32679-8' or numeroConta == '0000000054760' or numeroConta == '61951862-5'):
        return 'CC'
    elif numeroConta == '4984000000003458':
        return 'Visa'
    elif numeroConta == '32679-8/51':
        return 'CP'
    else:
        return ''

""""""""""""""""""""""""""""""""
""" Inserindo dados no EXCEL """
""""""""""""""""""""""""""""""""

def inserirTabelaExcel(row, arquivoXLSX, sheet):
    wb = load_workbook(arquivoXLSX)
    #print(wb.sheetnames)
    wb1 = wb[sheet]
    for x in range(0, len(row)):
        for y in range(0, len(row[x])):
            wb1.cell(row=x+1, column=y+1, value=row[x][y])

    dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)
    #dv.error('Você inseriu um valor que não está na lista de categorias')
    #dv.errorTitle('Entranda inválida')
    #dv.add('Fluxo_Caixa!G2:G8000')

    wb.save(arquivoXLSX)

def  retornarListaOperCSV(arquivo):
    with open(arquivo) as f:
        f_csv = csv.reader(f, lineterminator='\n', delimiter=';')
        linhas = []
        cabecalho = next(f_csv)
        print(cabecalho)
        linhas.append(cabecalho)
        for linha in f_csv:
            linhas.append(linha)

        return linhas